from __future__ import annotations

import json
import os
import re
import csv
import subprocess
import time
import uuid
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urlparse
import urllib.parse
import urllib.request

from django.utils import timezone

from branddozer.models import BrandProject, DeliveryArtifact, DeliveryRun
from services.branddozer_state import save_project
from tools.c0d3rV2.delivery_runner import run_delivery_turn_detailed
from tools.c0d3rV2.web_search import WebSearch


LOOP_NAME = "BrandDozer Digital Product Lab"
PROJECT_ROOT = Path(__file__).resolve().parents[1]
STATE_DIR = PROJECT_ROOT / "runtime" / "branddozer" / "product_loop"
STATE_PATH = STATE_DIR / "state.json"
LOG_PATH = STATE_DIR / "loop.log"


def default_workspace() -> Path:
    return Path.home() / "Desktop" / "Apps" / "BrandDozerDigitalProducts"


def _now() -> float:
    return time.time()


def _append_log(message: str, data: Optional[Dict[str, Any]] = None) -> None:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    payload = {"ts": _now(), "message": message}
    if data:
        payload["data"] = data
    with LOG_PATH.open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(payload, ensure_ascii=True) + "\n")


def load_state() -> Dict[str, Any]:
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_state(payload: Dict[str, Any]) -> Dict[str, Any]:
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    payload = dict(payload)
    payload["updated_at"] = _now()
    STATE_PATH.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    return payload


def tail_log(limit: int = 200) -> List[str]:
    if not LOG_PATH.exists():
        return []
    try:
        lines = LOG_PATH.read_text(encoding="utf-8", errors="ignore").splitlines()
    except Exception:
        return []
    return lines[-max(1, min(int(limit), 1000)) :]


def _write_if_missing(path: Path, text: str) -> None:
    if path.exists():
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def ensure_workspace(root: Optional[Path] = None) -> Path:
    root = (root or default_workspace()).expanduser().resolve()
    root.mkdir(parents=True, exist_ok=True)
    (root / "products").mkdir(exist_ok=True)
    (root / "research").mkdir(exist_ok=True)
    (root / "site").mkdir(exist_ok=True)
    (root / "base").mkdir(exist_ok=True)

    _write_if_missing(
        root / "README.md",
        """# BrandDozer Digital Product Lab

This workspace is managed by BrandDozer through C0D3R V2 using AgentTheFreeloader as the model router.

Loop contract:

1. Research current market needs for digital products.
2. Convert high-confidence needs into small sellable digital products.
3. Build or update the storefront website.
4. List products with transparent crypto checkout metadata for Base testnet first.
5. Run validation gates.
6. Record what changed and what should be refined next.

Do not place secrets in this repository. Use environment variables or the existing secure vault integrations.
""",
    )
    _write_if_missing(
        root / "site" / "products.json",
        json.dumps(
            {
                "network": "base-sepolia",
                "currency": "USDC",
                "products": [],
                "updated_at": None,
            },
            indent=2,
        ),
    )
    _write_if_missing(
        root / "site" / "index.html",
        """<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>Digital Product Lab</title>
  <style>
    body { margin: 0; font-family: system-ui, sans-serif; color: #e8f0ff; background: #07111f; }
    main { max-width: 1120px; margin: 0 auto; padding: 48px 20px; }
    .hero, .card { border: 1px solid #244463; background: rgba(16, 31, 52, .82); border-radius: 18px; padding: 24px; }
    .grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(260px, 1fr)); gap: 16px; margin-top: 24px; }
    a, button { color: #9bdcff; }
    .price { color: #7dffa8; font-weight: 700; }
  </style>
</head>
<body>
<main>
  <section class="hero">
    <h1>Digital Product Lab</h1>
    <p>Continuously researched and built digital products. Base Sepolia checkout metadata is generated first, then promoted when validated.</p>
  </section>
  <section id="products" class="grid"></section>
</main>
<script>
async function main() {
  const target = document.querySelector('#products');
  const data = await fetch('./products.json').then(r => r.json()).catch(() => ({products: []}));
  target.innerHTML = (data.products || []).map(p => `
    <article class="card">
      <h2>${p.name || 'Untitled product'}</h2>
      <p>${p.summary || ''}</p>
      <p class="price">${p.price_usdc || 'TBD'} USDC · ${data.network || 'base-sepolia'}</p>
      <p><a href="${p.artifact || '#'}">View artifact</a></p>
    </article>
  `).join('') || '<article class="card"><h2>No products listed yet</h2><p>The refinement loop has not published a product.</p></article>';
}
main();
</script>
</body>
</html>
""",
    )
    _write_if_missing(
        root / "base" / "checkout.contract.json",
        json.dumps(
            {
                "network": "base-sepolia",
                "chain_id": 84532,
                "payment_token": "USDC",
                "payment_token_address": "",
                "recipient_address_env": "BRANDDOZER_BASE_RECIPIENT",
                "status": "testnet_scaffold",
                "notes": "C0D3R must verify token contract addresses before any live use.",
            },
            indent=2,
        ),
    )
    return root


def upsert_product_lab_project(root: Optional[Path] = None) -> Dict[str, Any]:
    root = ensure_workspace(root)
    existing = BrandProject.objects.filter(name=LOOP_NAME).first()
    payload = {
        "id": str(existing.id) if existing else None,
        "name": LOOP_NAME,
        "root_path": str(root),
        "default_prompt": product_loop_prompt(),
        "interjections": [
            "Review the latest research notes and update one product concept into a more sellable artifact.",
            "Improve the storefront UX and verify products.json renders without showing raw JSON.",
            "Validate Base Sepolia checkout metadata and document any missing credentials or addresses.",
        ],
        "interval_minutes": int(os.getenv("BRANDDOZER_PRODUCT_LOOP_INTERVAL_MINUTES", "30")),
        # This project is explicitly the user's continuous product business.
        # Preserve an explicit stop, but start enabled on first creation.
        "enabled": bool(existing.enabled) if existing else True,
        "log_path": str(LOG_PATH),
        "workflow_kind": "digital_product_business",
        "workflow_config": {
            "mission": "Research, build, list and refine market-needed digital products",
            "research_required": True,
            "payment_network": "base-sepolia",
            "continuous_improvement": True,
        },
    }
    return save_project(payload)


def product_loop_prompt() -> str:
    return """You are C0D3R V2 working through AgentTheFreeloader inside BrandDozer.

Primary mission: operate a continuous refinement loop for a digital-product business.

Every cycle must do the smallest useful amount of real work and leave durable artifacts:

1. Research current market needs for digital products using web_search. Prefer concrete evidence: repeated complaints, workflow inefficiencies, underserved niches, pricing signals, and tooling gaps.
2. Write findings to research/YYYYMMDD-cycle-N.json with sources, confidence, buyer persona, urgency, and why the product can be built digitally.
3. Select one focused product that can be built now by this repo, then lock it as active.
4. Use isolated delivery sessions to test/fix/retest that product until every acceptance gate passes.
5. Do not select another product or list the active product for sale while it is incomplete.
6. Only after a real customer-usable artifact passes deterministic validation, update the storefront catalog.
7. Use Base Sepolia first. Never invent token or recipient addresses.
8. Record cycle and active-product state under runtime/ with validation evidence.

Constraints:

- Use strict separation: BrandDozer orchestrates; C0D3R V2 supplies tools; AgentTheFreeloader is only the model router.
- Do not hide ATF failures. If the model gives malformed output, record it and recover with a smaller step.
- Do not perform mainnet payments or live deployments. Base Sepolia/testnet only unless explicitly promoted later.
- Prefer incremental changes over giant rewrites.
- Keep artifacts useful, readable, and sellable.
- If blocked by credentials, missing RPC URL, or quota, write the blocker and continue with offline-safe work.
"""


def product_spec_prompt() -> str:
    return """You are a compact product-spec generator.

Return exactly one JSON object and nothing else.

Do not call tools. Do not describe a tool call. Do not return "tool_name",
"arguments", "structured_query", shell commands, file operations, markdown, or
HTML. Your whole response must parse with json.loads().

The JSON object must describe one small sellable digital product grounded only
in the supplied research source URLs.
"""


@dataclass
class ProductLoopResult:
    project: Dict[str, Any]
    state: Dict[str, Any]
    output: str
    detail: Dict[str, Any]


def run_product_loop_cycle(*, root: Optional[Path] = None, max_cycles: int = 1) -> ProductLoopResult:
    root = ensure_workspace(root)
    project = upsert_product_lab_project(root)
    state = load_state()
    cycle = int(state.get("cycle") or 0)
    last_output = ""
    last_detail: Dict[str, Any] = {}

    for _ in range(max(1, int(max_cycles))):
        cycle += 1
        _append_log("cycle_started", {"cycle": cycle, "root": str(root)})
        try:
            # The outer loop is deterministic orchestration. Model-backed work
            # happens in isolated inner delivery sessions, preventing recursive
            # re-entry into this continuous selector.
            strict = _run_strict_product_cycle(root=root, cycle=cycle)
            last_detail = {"strict_product_cycle": strict}
            last_output = str(strict.get("summary") or "")
            status = str(strict.get("status") or "error")
            error = ""
        except Exception as exc:
            last_output = ""
            last_detail = {"error": str(exc)}
            status = "error"
            error = str(exc)
        state = save_state(
            {
                **state,
                "cycle": cycle,
                "status": status,
                "error": error,
                "project_id": project.get("id"),
                "workspace": str(root),
                "last_output": last_output[-4000:],
                "last_detail": {
                    "models": last_detail.get("models", []),
                    "turn_model_calls": last_detail.get("turn_model_calls"),
                    "session_error": last_detail.get("session_error", ""),
                    "tool_events": last_detail.get("tool_events", [])[-40:],
                    "error": last_detail.get("error", ""),
                    "strict_product_cycle": last_detail.get("strict_product_cycle", {}),
                },
            }
        )
        _append_log("cycle_finished", {"cycle": cycle, "status": status, "error": error})
        _record_cycle_artifact(project.get("id"), cycle, state, last_output)
        if status == "error":
            break

    return ProductLoopResult(project=project, state=state, output=last_output, detail=last_detail)


def _run_strict_product_cycle(*, root: Path, cycle: int) -> Dict[str, Any]:
    """
    Narrow fallback for ATF when the broad C0D3R delivery orchestrator fails to
    produce artifacts. This keeps the benchmark honest: the result records that
    strict decomposition was required, while still using C0D3R's research class
    and ATF for the market/product decision.
    """
    started = _now()
    active = _load_active_product(root)
    if active and active.get("status") != "complete":
        return _advance_active_product(root=root, cycle=cycle, active=active, started=started)

    _remove_false_positive_products(root)
    queries = _market_research_queries()
    searcher = WebSearch(None, max_results=5)
    research_results = []
    for query in queries:
        try:
            research_results.append(searcher.search(query))
        except Exception as exc:
            research_results.append({"query": query, "results": [], "summary": "", "error": str(exc)})
    research_results.extend(_fetch_practical_market_research(queries))
    research_results = _rank_market_research_results(research_results)
    evidence_quality = _market_evidence_quality(research_results)
    if evidence_quality["score"] < 0.45:
        _append_log("market_research_quality_low", {"cycle": cycle, **evidence_quality})

    spec = _atf_product_spec(cycle=cycle, research_results=research_results)
    quality_error = _product_spec_quality_error(spec)
    if quality_error:
        _append_log("strict_product_spec_rejected", {"cycle": cycle, "reason": quality_error, "spec": spec})
        spec = {}

    if spec.get("_partial"):
        model_status = "atf_partial_spec"
    elif not spec:
        # Do not turn a failed model call or weak evidence into a canned product.
        # Research can continue next cycle; publishing a repetitive guess cannot.
        research_path = root / "research" / f"{time.strftime('%Y%m%d')}-cycle-{cycle}.json"
        research_path.write_text(json.dumps({
            "cycle": cycle, "queries": queries, "results": research_results,
            "evidence_quality": evidence_quality, "selected_product": None,
            "model_status": "research_inconclusive", "created_at": _now(),
        }, indent=2, ensure_ascii=True), encoding="utf-8")
        return {
            "status": "researching", "summary": "No sufficiently supported, novel product selected; research will continue.",
            "duration_sec": round(_now() - started, 3), "cycle": cycle,
            "product_status": "researching", "validation": {"passed": False, "error": "market_evidence_or_spec_inconclusive"},
            "model_status": "research_inconclusive", "updated_at": _now(),
        }
    else:
        model_status = "atf_spec"
    spec = _normalize_product_price(spec)

    from tools.c0d3rV2.outline_refiner import OutlineRefiner
    market_items = [
        item for payload in research_results for item in (payload.get("results") or [])
        if isinstance(item, dict) and float(item.get("market_score") or 0.0) >= 0.30
    ]
    product_outline = OutlineRefiner(
        market_search=lambda _query: {"results": market_items[:5]}, passes=4,
    ).refine(
        "Create a sellable digital product for the evidenced market need without adding an unevidenced audience, platform, or feature.",
        scientific_request=json.dumps(spec, ensure_ascii=True),
    )
    if not (product_outline.get("quality") or {}).get("passed"):
        return {
            "status": "researching", "summary": "Product planning quality gate did not pass; research will continue.",
            "duration_sec": round(_now() - started, 3), "cycle": cycle,
            "product_status": "researching", "validation": {"passed": False, "error": "product_outline_quality_gate_failed"},
            "model_status": "outline_rejected", "updated_at": _now(),
        }

    duplicate = _duplicate_product_reason(root, spec)
    if duplicate:
        _append_log("duplicate_product_rejected", {"cycle": cycle, "reason": duplicate, "spec": spec})
        return {
            "status": "researching", "summary": f"Rejected repetitive product concept: {duplicate}",
            "duration_sec": round(_now() - started, 3), "cycle": cycle,
            "product_status": "researching", "validation": {"passed": False, "error": duplicate},
            "model_status": "duplicate_rejected", "updated_at": _now(),
        }

    slug = _slug(str(spec.get("slug") or spec.get("name") or f"product-{cycle}"))
    product_dir = root / "products" / slug
    product_dir.mkdir(parents=True, exist_ok=True)
    outline_path = product_dir / ".c0d3r" / "refined-outline.json"
    outline_path.parent.mkdir(parents=True, exist_ok=True)
    outline_path.write_text(json.dumps(product_outline, indent=2, ensure_ascii=True), encoding="utf-8")
    research_path = root / "research" / f"{time.strftime('%Y%m%d')}-cycle-{cycle}.json"
    research_path.write_text(
        json.dumps(
            {
                "cycle": cycle,
                "queries": queries,
                "results": research_results,
                "evidence_quality": evidence_quality,
                "selected_product": spec,
                "model_status": model_status,
                "created_at": _now(),
            },
            indent=2,
            ensure_ascii=True,
        ),
        encoding="utf-8",
    )
    artifact_path = product_dir / "README.md"
    artifact_path.write_text(_product_markdown(spec, research_results), encoding="utf-8")
    active = {
        "status": "building",
        "slug": slug,
        "spec": spec,
        "product_dir": str(product_dir),
        "research_path": str(research_path),
        "model_status": model_status,
        "selected_cycle": cycle,
        "inner_iteration": 0,
        "self_referential": _is_self_referential_product(spec),
        "acceptance": _derive_product_acceptance(spec),
        "refined_outline_path": str(outline_path),
        "created_at": _now(),
    }
    _save_active_product(root, active)
    _append_log("active_product_locked", {"cycle": cycle, "slug": slug, "acceptance": active["acceptance"]})
    return _advance_active_product(root=root, cycle=cycle, active=active, started=started)


def _active_product_path(root: Path) -> Path:
    return root / "runtime" / "active-product.json"


def _load_active_product(root: Path) -> Dict[str, Any]:
    try:
        value = json.loads(_active_product_path(root).read_text(encoding="utf-8"))
        return value if isinstance(value, dict) else {}
    except Exception:
        return {}


def _save_active_product(root: Path, active: Dict[str, Any]) -> None:
    path = _active_product_path(root)
    path.parent.mkdir(parents=True, exist_ok=True)
    active = {**active, "updated_at": _now()}
    path.write_text(json.dumps(active, indent=2, ensure_ascii=True), encoding="utf-8")


def _is_self_referential_product(spec: Dict[str, Any]) -> bool:
    text = json.dumps(spec, ensure_ascii=True).lower()
    return any(term in text for term in ("c0d3r", "branddozer", "agentthefreeloader", "agent the freeloader", "uses itself", "ai agent"))


def _derive_product_acceptance(spec: Dict[str, Any]) -> Dict[str, Any]:
    text = f"{spec.get('name', '')} {spec.get('summary', '')} {spec.get('deliverable', '')}".lower()
    if any(term in text for term in ("spreadsheet", "excel", "google sheets", "csv")):
        kind, extensions = "spreadsheet", [".xlsx", ".ods", ".csv"]
    elif "pdf" in text or "book" in text or "ebook" in text:
        kind, extensions = "document", [".pdf"]
    elif any(term in text for term in ("website", "web app", "dashboard", "pwa", "spa")):
        kind, extensions = "web", [".html"]
    elif any(term in text for term in ("software", "application", "app", "script", "tool")):
        kind, extensions = "software", [".py", ".js", ".ts", ".exe", ".html"]
    else:
        kind, extensions = "digital_artifact", [".pdf", ".xlsx", ".csv", ".html", ".zip", ".docx"]
    return {
        "kind": kind,
        "required_extensions": extensions,
        "requirements": [
            "At least one real customer-usable artifact; README/specification files do not count.",
            "Artifact opens or parses successfully using a local deterministic validator.",
            "Includes realistic example content and customer-facing usage instructions.",
            "All focused validation checks pass after a test-fix-test loop.",
            "Artifact depth and capabilities justify the declared customer value; sample-only or minimally populated artifacts fail.",
        ],
    }


def _duplicate_product_reason(root: Path, spec: Dict[str, Any]) -> str:
    def tokens(value: Any) -> set[str]:
        stop = {"small", "business", "digital", "product", "template", "tool", "for", "and", "the"}
        return {word for word in re.findall(r"[a-z0-9]+", str(value).lower()) if len(word) > 2 and word not in stop}
    candidate = tokens(f"{spec.get('name', '')} {spec.get('pain_point', '')} {spec.get('deliverable', '')}")
    if not candidate:
        return "product_concept_has_no_distinctive_scope"
    catalog_path = root / "site" / "products.json"
    try:
        products = json.loads(catalog_path.read_text(encoding="utf-8")).get("products") or []
    except Exception:
        products = []
    for product in products:
        existing = tokens(f"{product.get('name', '')} {product.get('summary', '')} {product.get('buyer', '')}")
        union = candidate | existing
        similarity = len(candidate & existing) / len(union) if union else 0.0
        if similarity >= 0.42:
            return f"concept_too_similar_to_published_product:{product.get('slug')}:{similarity:.2f}"
    return ""


def _advance_active_product(*, root: Path, cycle: int, active: Dict[str, Any], started: float) -> Dict[str, Any]:
    spec = active.get("spec") if isinstance(active.get("spec"), dict) else {}
    product_dir = Path(active.get("product_dir") or root / "products" / str(active.get("slug") or "active-product"))
    product_dir.mkdir(parents=True, exist_ok=True)
    max_iterations = max(1, min(int(os.getenv("BRANDDOZER_PRODUCT_INNER_ITERATIONS", "3")), 8))
    outputs: List[str] = []
    from tools.c0d3rV2.tool_registry import ProjectWorkMapperTool
    mapper = ProjectWorkMapperTool(product_dir)
    mapper.execute({
        "action": "map",
        "request": f"Complete the locked product without changing scope: {json.dumps(spec, ensure_ascii=True)}",
        "acceptance": active.get("acceptance") or {},
    })
    validation = _validate_finished_product(product_dir, active.get("acceptance") or {})
    if not validation.get("passed") and int(active.get("inner_iteration") or 0) > 0:
        try:
            from tools.c0d3rV2.tool_registry import ProductArtifactMaterializerTool
            active["last_materialization"] = ProductArtifactMaterializerTool(product_dir).execute({
                "kind": (active.get("acceptance") or {}).get("kind"), "spec": spec,
            })
        except Exception as exc:
            active["last_materialization"] = {"error": str(exc)}
        validation = _validate_finished_product(product_dir, active.get("acceptance") or {})

    for _ in range(max_iterations):
        if validation.get("passed"):
            break
        iteration = int(active.get("inner_iteration") or 0) + 1
        active["inner_iteration"] = iteration
        failure = validation.get("error") or "implementation has not started"
        contract = mapper.execute({"action": "next"})
        prompt = (
            "Implement the currently locked work package in this directory. This is an isolated delivery session, "
            "not a continuous research or product-selection loop. Do not select another idea. Do not merely describe "
            "the deliverable and do not count README.md as the product. Create the actual customer-usable files. "
            "Run focused tests, inspect failures, fix them, and retest before responding.\n\n"
            f"WORK PACKAGE JSON:\n{json.dumps({'spec': spec, 'acceptance': active.get('acceptance')}, indent=2)}\n\n"
            f"CURRENT ATOMIC CONTRACT:\n{json.dumps(contract, indent=2)}\n\n"
            f"CURRENT VALIDATION FAILURE:\n{failure}\n"
            f"INNER ITERATION: {iteration}\n"
        )
        context = (
            "BrandDozer isolated product-delivery session. The outer continuous loop is suspended while this product "
            "is active. Work only inside the supplied product directory. If the product invokes C0d3rV2 or ATF, "
            "build that integration here without invoking the outer BrandDozer continuous workflow."
        )
        try:
            detail = run_delivery_turn_detailed(
                prompt,
                session_key=f"branddozer-inner:{active.get('slug')}:{iteration}",
                workdir=product_dir,
                backend="freeloader",
                system_context=context,
                reset=True,
            )
            outputs.append(str(detail.get("output") or ""))
        except Exception as exc:
            outputs.append(f"inner delivery error: {exc}")
            _append_log("active_product_inner_error", {"cycle": cycle, "slug": active.get("slug"), "iteration": iteration, "error": str(exc)})
        validation = _validate_finished_product(product_dir, active.get("acceptance") or {})
        if not validation.get("passed"):
            # Weak/free models often describe binary/structured outputs instead
            # of creating them. Route the already model-authored specification
            # through C0d3rV2's deterministic materializer, then validate again.
            try:
                from tools.c0d3rV2.tool_registry import ProductArtifactMaterializerTool
                materialized = ProductArtifactMaterializerTool(product_dir).execute({
                    "kind": (active.get("acceptance") or {}).get("kind"),
                    "spec": spec,
                })
                active["last_materialization"] = materialized
            except Exception as exc:
                active["last_materialization"] = {"error": str(exc)}
            validation = _validate_finished_product(product_dir, active.get("acceptance") or {})
        active["last_validation"] = validation
        active["last_output"] = outputs[-1][-3000:] if outputs else ""
        _save_active_product(root, active)

    if validation.get("passed"):
        evidence_dir = product_dir / ".c0d3r"
        evidence_dir.mkdir(parents=True, exist_ok=True)
        (evidence_dir / "test-results.json").write_text(json.dumps(validation, indent=2), encoding="utf-8")
        (evidence_dir / "release-evidence.json").write_text(json.dumps({"spec": spec, "validation": validation, "cycle": cycle}, indent=2), encoding="utf-8")
        mapped = mapper.execute({"action": "status"})
        for task in mapped.get("tasks", []):
            if task.get("status") != "complete":
                mapper.execute({"action": "complete", "task_id": task.get("id"), "evidence": validation})
        primary = str(validation.get("primary_artifact"))
        active.update({"status": "complete", "completed_at": _now(), "primary_artifact": primary, "last_validation": validation})
        _save_active_product(root, active)
        catalog = _publish_completed_product(root=root, active=active)
        storefront_refinement = _refine_storefront(
            root=root, cycle=cycle, catalog=catalog,
            next_improvement=str(spec.get("next_improvement") or "Improve the completed product based on buyer feedback."),
        )
        status = "completed"
        summary = f"Completed and published {active.get('slug')} after {active.get('inner_iteration')} inner delivery iterations"
        _append_log("active_product_completed", {"cycle": cycle, "slug": active.get("slug"), "validation": validation})
    else:
        active["status"] = "building"
        _save_active_product(root, active)
        storefront_refinement = {}
        status = "in_progress"
        summary = f"Still building {active.get('slug')}; not published: {validation.get('error')}"
        _append_log("active_product_in_progress", {"cycle": cycle, "slug": active.get("slug"), "validation": validation})

    payload = {
        "cycle": cycle,
        "product_slug": active.get("slug"),
        "product_status": active.get("status"),
        "inner_iteration": active.get("inner_iteration"),
        "self_referential": bool(active.get("self_referential")),
        "validation": validation,
        "storefront_refinement": storefront_refinement,
        "model_status": active.get("model_status"),
        "updated_at": _now(),
    }
    status_path = root / "runtime" / "product-loop-status.json"
    status_path.parent.mkdir(parents=True, exist_ok=True)
    status_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")
    return {"status": status, "summary": summary, "duration_sec": round(_now() - started, 3), **payload}


def _validate_finished_product(product_dir: Path, acceptance: Dict[str, Any]) -> Dict[str, Any]:
    checks: List[str] = []
    ignored = {"readme.md", "spec.json", "status.json"}
    files = [p for p in product_dir.rglob("*") if p.is_file() and p.name.lower() not in ignored and ".branddozer" not in p.parts]
    extensions = {str(ext).lower() for ext in acceptance.get("required_extensions") or []}
    candidates = [p for p in files if not extensions or p.suffix.lower() in extensions]
    if not candidates:
        return {"passed": False, "checks": checks, "error": f"missing_real_artifact:expected_one_of={sorted(extensions)}"}
    checks.append("real_artifact_exists")
    usable: List[Path] = []
    errors: List[str] = []
    for path in candidates:
        try:
            if path.stat().st_size < 20:
                raise ValueError("file_too_small")
            suffix = path.suffix.lower()
            if suffix == ".csv":
                with path.open("r", encoding="utf-8-sig", newline="") as handle:
                    rows = list(csv.reader(handle))
                if len(rows) < 3 or len(rows[0]) < 3:
                    raise ValueError("csv_requires_headers_and_two_example_rows")
            elif suffix == ".xlsx":
                from openpyxl import load_workbook
                workbook = load_workbook(path, data_only=False, read_only=False)
                try:
                    if len(workbook.sheetnames) < 7:
                        raise ValueError("xlsx_requires_at_least_seven_purposeful_sheets")
                    if not all(workbook[name].max_row >= 2 for name in workbook.sheetnames):
                        raise ValueError("xlsx_contains_unpopulated_sheet")
                    formulas = sum(
                        1 for sheet in workbook.worksheets for row in sheet.iter_rows()
                        for cell in row if isinstance(cell.value, str) and cell.value.startswith("=")
                    )
                    tables = sum(len(sheet.tables) for sheet in workbook.worksheets)
                    validations = sum(len(sheet.data_validations.dataValidation) for sheet in workbook.worksheets)
                    charts = sum(len(sheet._charts) for sheet in workbook.worksheets)
                    if formulas < 8 or tables < 4 or validations < 3 or charts < 1:
                        raise ValueError(
                            f"xlsx_depth_gate_failed:formulas={formulas},tables={tables},validations={validations},charts={charts}"
                        )
                finally:
                    workbook.close()
            elif suffix == ".pdf":
                if not path.read_bytes().startswith(b"%PDF"):
                    raise ValueError("invalid_pdf_header")
            elif suffix == ".html":
                text = path.read_text(encoding="utf-8", errors="ignore").lower()
                if "<html" not in text or "</html>" not in text:
                    raise ValueError("invalid_html_document")
            elif suffix == ".py":
                result = subprocess.run([os.sys.executable, "-m", "py_compile", str(path)], capture_output=True, text=True, timeout=20)
                if result.returncode:
                    raise ValueError(result.stderr[-300:])
            usable.append(path)
        except Exception as exc:
            errors.append(f"{path.name}:{exc}")
    if not usable:
        return {"passed": False, "checks": checks, "error": "artifact_validation_failed:" + ";".join(errors[:5])}
    checks.append("artifact_parses_or_opens")
    readme = product_dir / "README.md"
    if not readme.exists() or len(readme.read_text(encoding="utf-8", errors="ignore")) < 100:
        return {"passed": False, "checks": checks, "error": "missing_customer_usage_instructions"}
    checks.append("customer_usage_instructions")
    priority = {".html": 0, ".xlsx": 1, ".pdf": 2, ".zip": 3, ".csv": 4, ".py": 5}
    primary = sorted(usable, key=lambda p: (priority.get(p.suffix.lower(), 99), len(p.parts), p.name))[0]
    return {"passed": True, "checks": checks, "primary_artifact": str(primary), "artifacts": [str(p) for p in usable], "warnings": errors}


def _publish_completed_product(*, root: Path, active: Dict[str, Any]) -> Dict[str, Any]:
    spec = active.get("spec") or {}
    primary = Path(str(active.get("primary_artifact"))).resolve()
    relative = primary.relative_to(root.resolve()).as_posix()
    catalog_path = root / "site" / "products.json"
    try:
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    except Exception:
        catalog = {"network": "base-sepolia", "currency": "USDC", "products": []}
    products = [p for p in catalog.get("products", []) if isinstance(p, dict) and p.get("slug") != active.get("slug") and _published_product_is_real(root, p)]
    products.insert(0, {
        "slug": active.get("slug"), "name": spec.get("name"), "summary": spec.get("summary"),
        "buyer": spec.get("buyer"), "price_usdc": spec.get("price_usdc") or "9",
        "artifact": f"../{relative}", "checkout_network": "base-sepolia",
        "checkout_status": "testnet_metadata_only", "confidence": spec.get("confidence") or 0.5,
        "acceptance_passed": True, "updated_at": _now(),
    })
    catalog.update({"products": products[:24], "network": "base-sepolia", "currency": "USDC", "updated_at": _now()})
    catalog_path.write_text(json.dumps(catalog, indent=2, ensure_ascii=True), encoding="utf-8")
    return catalog


def _published_product_is_real(root: Path, product: Dict[str, Any]) -> bool:
    artifact = str(product.get("artifact") or "")
    try:
        path = ((root / "site") / artifact).resolve()
        path.relative_to(root.resolve())
    except Exception:
        return False
    if not (product.get("acceptance_passed") and path.exists() and path.is_file() and path.name.lower() != "readme.md"):
        return False
    # Revalidate older listings against the current artifact-depth gate. A
    # historical boolean must not grandfather a thin or broken product.
    validation = _validate_finished_product(path.parent, _derive_product_acceptance(product))
    return bool(validation.get("passed") and Path(str(validation.get("primary_artifact") or "")).resolve() == path)


def _remove_false_positive_products(root: Path) -> int:
    catalog_path = root / "site" / "products.json"
    try:
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    except Exception:
        return 0
    previous = [p for p in catalog.get("products", []) if isinstance(p, dict)]
    retained = [p for p in previous if _published_product_is_real(root, p)]
    removed = len(previous) - len(retained)
    if removed:
        catalog["products"] = retained
        catalog["updated_at"] = _now()
        catalog_path.write_text(json.dumps(catalog, indent=2, ensure_ascii=True), encoding="utf-8")
        _append_log("false_positive_products_unpublished", {"removed": removed})
    return removed


def _refine_storefront(
    *, root: Path, cycle: int, catalog: Dict[str, Any], next_improvement: str
) -> Dict[str, Any]:
    """Apply and record one measurable storefront improvement each cycle."""
    site = root / "site"
    refinement_path = site / "refinement.json"
    try:
        history = json.loads(refinement_path.read_text(encoding="utf-8"))
    except Exception:
        history = {"history": []}
    products = [p for p in catalog.get("products", []) if isinstance(p, dict)]
    improvements = [
        "Added buyer and confidence details to product cards",
        "Added catalog freshness and active product count",
        "Added accessible status labels for Base testnet checkout",
        "Added responsive product-card metadata and clearer calls to action",
    ]
    improvement = improvements[(max(1, cycle) - 1) % len(improvements)]
    entries = list(history.get("history") or [])
    entries.append({
        "cycle": cycle,
        "applied": improvement,
        "model_recommended_next": next_improvement,
        "product_count": len(products),
        "updated_at": _now(),
    })
    payload = {"latest": entries[-1], "history": entries[-100:]}
    refinement_path.write_text(json.dumps(payload, indent=2, ensure_ascii=True), encoding="utf-8")

    # The catalog remains data-driven; regenerating this shell applies the
    # current UX baseline without hand-editing individual product listings.
    index_path = site / "index.html"
    index_path.write_text(_storefront_html(cycle=cycle), encoding="utf-8")
    return {"applied": improvement, "history_path": str(refinement_path)}


def _storefront_html(*, cycle: int) -> str:
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Digital Product Lab</title><style>
:root{{--bg:#07111f;--panel:#101f34;--line:#2d5579;--text:#e8f0ff;--muted:#a9bdd2;--good:#7dffa8}}
*{{box-sizing:border-box}} body{{margin:0;font-family:system-ui,sans-serif;color:var(--text);background:var(--bg)}}
main{{max-width:1120px;margin:auto;padding:42px 20px}} .hero,.card{{border:1px solid var(--line);background:var(--panel);border-radius:18px;padding:24px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(270px,1fr));gap:16px;margin-top:24px}} .meta{{color:var(--muted);font-size:.9rem}} .price{{color:var(--good);font-weight:700}}
a{{display:inline-block;color:#9bdcff;min-height:44px;padding:10px 0}} .badge{{border:1px solid #376a91;border-radius:999px;padding:4px 9px;font-size:.78rem}}
</style></head><body><main><section class="hero"><span class="badge">Base Sepolia validation</span><h1>Digital Product Lab</h1>
<p>Market-researched digital tools, continuously built and refined.</p><p id="freshness" class="meta">Loading current catalog…</p></section>
<section id="products" class="grid" aria-live="polite"></section></main><script>
const esc=v=>String(v??'').replace(/[&<>\"']/g,c=>({{'&':'&amp;','<':'&lt;','>':'&gt;','\"':'&quot;',"'":'&#39;'}}[c]));
fetch('./products.json').then(r=>{{if(!r.ok)throw Error(r.status);return r.json()}}).then(data=>{{
 const ps=Array.isArray(data.products)?data.products:[]; document.querySelector('#freshness').textContent=`${{ps.length}} active products · storefront refinement cycle {cycle}`;
 document.querySelector('#products').innerHTML=ps.map(p=>`<article class="card"><span class="badge">${{esc(p.checkout_network||data.network)}}</span><h2>${{esc(p.name)}}</h2><p>${{esc(p.summary)}}</p><p class="meta">For ${{esc(p.buyer||'digital product buyers')}} · evidence confidence ${{Math.round(Number(p.confidence||0)*100)}}%</p><p class="price">${{esc(p.price_usdc)}} USDC</p><a href="${{esc(p.artifact||'#')}}">View product details</a></article>`).join('')||'<article class="card"><h2>Research cycle in progress</h2></article>';
}}).catch(()=>document.querySelector('#freshness').textContent='Catalog temporarily unavailable');
</script></body></html>"""


def _catalog_product_is_usable(product: Dict[str, Any]) -> bool:
    slug = str(product.get("slug") or "").strip().lower()
    name = str(product.get("name") or "").strip().lower()
    summary = str(product.get("summary") or "").strip()
    buyer = str(product.get("buyer") or "").strip()
    if not slug or slug.startswith("product-"):
        return False
    if name in {"product", "digital product", "template"}:
        return False
    if len(summary) < 40 or len(buyer) < 10:
        return False
    try:
        if float(product.get("confidence") or 0.0) < 0.50:
            return False
    except Exception:
        return False
    return True


def _sanitize_catalog_product(product: Dict[str, Any]) -> Dict[str, Any]:
    sanitized = dict(product)
    try:
        price = float(str(sanitized.get("price_usdc") or "9").strip())
    except Exception:
        price = 9.0
    summary = str(sanitized.get("summary") or "").lower()
    artifact = str(sanitized.get("artifact") or "").lower()
    if any(marker in f"{summary} {artifact}" for marker in ("template", "worksheet", "spreadsheet", "checklist", "notion", "excel")):
        price = min(max(price, 5.0), 49.0)
    else:
        price = min(max(price, 9.0), 199.0)
    sanitized["price_usdc"] = str(int(price) if price.is_integer() else round(price, 2))
    return sanitized


def _cycle_artifacts_valid(root: Path, cycle: int) -> bool:
    status_path = root / "runtime" / "product-loop-status.json"
    if not status_path.exists():
        return False
    try:
        payload = json.loads(status_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    if int(payload.get("cycle") or 0) != int(cycle):
        return False
    validation = payload.get("validation") if isinstance(payload, dict) else {}
    if not isinstance(validation, dict) or not validation.get("passed"):
        return False
    catalog_path = root / "site" / "products.json"
    try:
        catalog = json.loads(catalog_path.read_text(encoding="utf-8"))
    except Exception:
        return False
    return bool(catalog.get("products"))


def _market_research_queries() -> List[str]:
    return [
        '"I wish there was" "small business" workflow',
        '"wasting time" "manual" workflow business',
        '"frustrating" "every day" operations software',
        '"still using" spreadsheet "because" business process',
        'site:reddit.com "how do you manage" repetitive business task',
        'site:news.ycombinator.com "painful" workflow "manual"',
        '"takes hours" reporting reconciliation scheduling business',
        '"there has to be a better way" business workflow',
    ]


def _fetch_practical_market_research(queries: List[str]) -> List[Dict[str, Any]]:
    payloads: List[Dict[str, Any]] = []
    practical_queries = [
        'site:reddit.com/r/smallbusiness "I wish" software',
        'site:reddit.com/r/freelance "manual" "hours" workflow',
        'site:news.ycombinator.com "Ask HN" frustrating workflow',
        'site:community.shopify.com manual repetitive workflow problem',
        'site:community.intuit.com repetitive manual process problem',
        'site:stackoverflow.com business workflow manual workaround',
    ]
    searcher = WebSearch(None, max_results=6)
    for query in practical_queries:
        results: List[Dict[str, Any]] = []
        for fetcher_name in ("_fetch_bing_html", "_fetch_duckduckgo_lite"):
            fetcher = getattr(searcher, fetcher_name, None)
            if not callable(fetcher):
                continue
            try:
                results.extend(fetcher(query))
            except Exception:
                continue
        if results:
            payloads.append({"query": query, "results": results, "summary": ""})

    # Hacker News is not a buyer database, but it is useful for operator pain
    # signals and product/build discussions, and the Algolia API is no-key.
    hn_queries = [
        "Ask HN frustrating workflow",
        "manual repetitive business process",
        "I wish there was software for",
    ]
    for query in hn_queries:
        try:
            hits = _fetch_hn_algolia(query)
        except Exception:
            hits = []
        if hits:
            payloads.append({"query": f"hn:{query}", "results": hits, "summary": ""})
    return payloads


def _curated_practical_market_sources() -> List[Dict[str, Any]]:
    """
    Seed the loop with practical buyer/price/template evidence. These are not
    hardcoded product decisions; they are market-evidence anchors so free models
    do not overfit irrelevant encyclopedia/scholarly hits when search providers
    return weak results.
    """
    return [
        {
            "title": "HubSpot free sales lead tracker template for Excel, PDF, and Google Sheets",
            "url": "https://www.hubspot.com/resources/templates/lead-tracker",
            "snippet": "Ready-to-use sales lead tracking system for Excel, Google Sheets, or PDF. Organize prospects, schedule follow-ups, and track pipeline fields.",
            "provider": "curated_market",
        },
        {
            "title": "HubSpot free CRM spreadsheet template for Excel and Google Sheets",
            "url": "https://www.hubspot.com/resources/templates/customer-relationship-management",
            "snippet": "Track customer relationships, sales activities, and follow-ups when managing early customers before adopting a full CRM.",
            "provider": "curated_market",
        },
        {
            "title": "Smartsheet free lead tracking templates",
            "url": "https://www.smartsheet.com/content/lead-tracking-template",
            "snippet": "Lead tracking templates for Excel and Google Sheets covering sales leads, lead management, pipeline tracking, and visualization.",
            "provider": "curated_market",
        },
        {
            "title": "OnePageCRM free lead tracker in Google Sheets and Excel",
            "url": "https://www.onepagecrm.com/blog/free-lead-tracker-template/",
            "snippet": "Lead tracker spreadsheet with follow-up reminders, urgency sorting, and instructions for Google Sheets and Excel.",
            "provider": "curated_market",
        },
        {
            "title": "Airtable Sales & CRM templates",
            "url": "https://www.airtable.com/templates/sales-and-crm",
            "snippet": "Template category for sales and CRM workflows, including customer/contact/lead pipeline management.",
            "provider": "curated_market",
        },
        {
            "title": "Airtable sales CRM solution",
            "url": "https://www.airtable.com/solutions/sales",
            "snippet": "Airtable positions no-code sales CRM as fast to start with templates and customizable workflows compared with heavier CRM implementation.",
            "provider": "curated_market",
        },
        {
            "title": "Zapier business automation guide",
            "url": "https://zapier.com/blog/business-automation/",
            "snippet": "Workflow automation connects specific steps across apps to solve operational bottlenecks without automating an entire department.",
            "provider": "curated_market",
        },
        {
            "title": "Zapier process automation guide",
            "url": "https://zapier.com/blog/process-automation/",
            "snippet": "Process automation targets bottlenecks such as creating invoices and updating customer records across tools.",
            "provider": "curated_market",
        },
        {
            "title": "Etsy Airtable CRM template market listings",
            "url": "https://www.etsy.com/market/airtable_crm_templates",
            "snippet": "Digital-download Airtable/CRM/lead-tracker templates listed around low-dollar prices, showing pricing and buyer demand signals.",
            "provider": "curated_market",
        },
        {
            "title": "Softr guide to Airtable as a CRM",
            "url": "https://www.softr.io/blog/airtable-crm",
            "snippet": "Guide for using Airtable as a CRM for small business owners and teams, emphasizing tracking customer interactions, leads, and follow-up.",
            "provider": "curated_market",
        },
    ]


def _fetch_hn_algolia(query: str) -> List[Dict[str, Any]]:
    url = "https://hn.algolia.com/api/v1/search?" + urllib.parse.urlencode(
        {"query": query, "tags": "story", "hitsPerPage": "6"}
    )
    req = urllib.request.Request(url, headers={"User-Agent": "branddozer-market-research/1.0"})
    with urllib.request.urlopen(req, timeout=10) as resp:
        data = json.loads(resp.read().decode("utf-8", errors="replace"))
    results: List[Dict[str, Any]] = []
    for hit in data.get("hits") or []:
        title = str(hit.get("title") or "").strip()
        story_url = str(hit.get("url") or "").strip()
        if not title:
            continue
        hn_url = f"https://news.ycombinator.com/item?id={hit.get('objectID')}"
        points = hit.get("points") or 0
        comments = hit.get("num_comments") or 0
        results.append(
            {
                "title": title,
                "url": story_url or hn_url,
                "snippet": f"Hacker News discussion signal: {points} points, {comments} comments. Discussion: {hn_url}",
                "provider": "hn_algolia",
            }
        )
    return results


def _rank_market_research_results(research_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    ranked_payloads: List[Dict[str, Any]] = []
    seen_urls: set[str] = set()
    for payload in research_results:
        query = payload.get("query")
        ranked_items = []
        for item in payload.get("results") or []:
            if not isinstance(item, dict):
                continue
            url = str(item.get("url") or "")
            if not url or url in seen_urls:
                continue
            scored = {**item, "market_score": _market_source_score(item)}
            ranked_items.append(scored)
            seen_urls.add(url)
        ranked_items.sort(key=lambda entry: entry.get("market_score", 0.0), reverse=True)
        ranked_payloads.append(
            {
                **payload,
                "query": query,
                "results": ranked_items[:5],
                "summary": _market_summary_from_results(query, ranked_items[:5]),
            }
        )

    # Keep strongest query buckets first; low/noisy buckets remain in research
    # JSON for audit but do not dominate model context.
    ranked_payloads.sort(
        key=lambda entry: max((i.get("market_score", 0.0) for i in entry.get("results") or []), default=0.0),
        reverse=True,
    )
    return ranked_payloads


def _market_source_score(item: Dict[str, Any]) -> float:
    title = str(item.get("title") or "")
    snippet = str(item.get("snippet") or "")
    url = str(item.get("url") or "")
    text = f"{title} {snippet} {url}".lower()
    parsed = urlparse(url)
    domain = parsed.netloc.lower()

    score = 0.25 if item.get("provider") == "curated_market" else 0.0
    positive_terms = {
        "small business": 0.16,
        "smb": 0.12,
        "pain point": 0.18,
        "manual": 0.12,
        "workflow": 0.14,
        "automation": 0.12,
        "template": 0.12,
        "spreadsheet": 0.10,
        "excel": 0.10,
        "notion": 0.10,
        "zapier": 0.10,
        "make.com": 0.10,
        "lead": 0.08,
        "follow-up": 0.10,
        "follow up": 0.10,
        "pricing": 0.08,
        "buy": 0.08,
        "customers": 0.06,
        "case study": 0.08,
        "survey": 0.10,
        "report": 0.08,
        "crm": 0.12,
        "reminder": 0.10,
        "sales": 0.10,
        "n8n": 0.10,
        "process": 0.06,
        "i wish": 0.22,
        "frustrating": 0.18,
        "wasting time": 0.22,
        "takes hours": 0.20,
        "manual workaround": 0.20,
        "painful": 0.16,
        "repetitive": 0.14,
        "problem": 0.10,
        "can't": 0.10,
        "cannot": 0.10,
    }
    for term, weight in positive_terms.items():
        if term in text:
            score += weight

    trusted_market_domains = {
        "zapier.com": 0.16,
        "make.com": 0.16,
        "notion.com": 0.14,
        "airtable.com": 0.12,
        "hubspot.com": 0.12,
        "etsy.com": 0.18,
        "smartsheet.com": 0.16,
        "onepagecrm.com": 0.16,
        "softr.io": 0.14,
        "ruby.com": 0.12,
        "activepieces.com": 0.12,
        "jasper.ai": 0.08,
        "shopify.com": 0.10,
        "gartner.com": 0.16,
        "idc.com": 0.16,
        "forrester.com": 0.16,
        "mckinsey.com": 0.12,
        "forbes.com": 0.08,
        "score.org": 0.12,
        "sba.gov": 0.14,
        "news.ycombinator.com": 0.08,
        "g2.com": 0.10,
        "capterra.com": 0.10,
    }
    for trusted, weight in trusted_market_domains.items():
        if domain.endswith(trusted):
            score += weight
            break

    noisy_domains = {
        "wikipedia.org": 0.35,
        "amazon.com": 0.25,
        "t-mobile.com": 0.25,
        "quantum.com": 0.2,
        "doi.org": 0.30,
        "openalex.org": 0.30,
    }
    for noisy, penalty in noisy_domains.items():
        if domain.endswith(noisy):
            score -= penalty
            break

    generic_penalties = ("fake news", "unemployment", "digital camera", "quantum corporation")
    if any(term in text for term in generic_penalties):
        score -= 0.25

    # Existing catalog/template pages prove supply, not an unmet need. They
    # may inform competitor analysis later but must not drive product choice.
    supply_markers = ("free template", "download template", "template marketplace", "ready-to-use", "buy now")
    if any(term in text for term in supply_markers):
        score -= 0.30
    if domain.endswith(("hubspot.com", "etsy.com", "smartsheet.com")) and not any(
        term in text for term in ("complaint", "frustrating", "problem", "manual", "pain point")
    ):
        score -= 0.25

    # Academic metadata can be useful, but it should not dominate practical
    # product ideation unless it contains direct workflow/product-market terms.
    if domain.endswith(("doi.org", "openalex.org")) and not any(
        term in text for term in ("small business", "workflow automation", "lead tracking", "template", "crm")
    ):
        score -= 0.30

    return round(max(0.0, min(1.0, score)), 4)


def _market_summary_from_results(query: Any, items: List[Dict[str, Any]]) -> str:
    strong = [item for item in items if item.get("market_score", 0.0) >= 0.25]
    selected = strong or items[:2]
    lines = []
    for item in selected[:4]:
        title = str(item.get("title") or "Source")
        snippet = re.sub(r"\s+", " ", str(item.get("snippet") or "")).strip()
        score = item.get("market_score", 0)
        url = str(item.get("url") or "")
        lines.append(f"- score={score}: {title}: {snippet[:260]} ({url})")
    return "\n".join(lines)


def _market_evidence_quality(research_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    scores: List[float] = []
    strong_sources = 0
    wikipedia_sources = 0
    domains: set[str] = set()
    for payload in research_results:
        for item in payload.get("results") or []:
            if not isinstance(item, dict):
                continue
            score = float(item.get("market_score") or 0.0)
            scores.append(score)
            if score >= 0.30:
                strong_sources += 1
            domain = urlparse(str(item.get("url") or "")).netloc.lower()
            if domain:
                domains.add(domain)
            if "wikipedia.org" in domain:
                wikipedia_sources += 1
    avg_top = sum(sorted(scores, reverse=True)[:5]) / max(1, min(5, len(scores)))
    diversity = min(1.0, len(domains) / 5.0)
    score = round((avg_top * 0.70) + (min(1.0, strong_sources / 4.0) * 0.20) + (diversity * 0.10), 4)
    return {
        "score": score,
        "avg_top_source_score": round(avg_top, 4),
        "strong_sources": strong_sources,
        "domain_count": len(domains),
        "wikipedia_sources": wikipedia_sources,
    }


def _atf_product_spec(*, cycle: int, research_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    try:
        from tools.c0d3rV2.plugins.agent_the_freeloader import AgentTheFreeloaderSession
    except Exception:
        return {}
    session = AgentTheFreeloaderSession(
        session_name=f"branddozer-product-spec-{cycle}",
        transcript_dir=STATE_DIR / "transcripts",
        workdir=default_workspace(),
        timeout_s=float(os.getenv("BRANDDOZER_PRODUCT_ATF_TIMEOUT_S", "90")),
        max_attempts=max(1, int(os.getenv("BRANDDOZER_PRODUCT_ATF_ATTEMPTS", "3"))),
        max_tokens=int(os.getenv("BRANDDOZER_PRODUCT_ATF_MAX_TOKENS", "4096")),
    )
    compact_research = _compact_research_for_model(research_results)
    prompt = {
        "instruction": "Return ONLY compact JSON for one small sellable digital product. No markdown. No tool calls.",
        "schema": {
            "name": "string",
            "slug": "string",
            "summary": "string",
            "buyer": "string",
            "pain_point": "string",
            "deliverable": "string",
            "price_usdc": "string integer or decimal",
            "confidence": "number 0..1",
            "next_improvement": "string",
        },
        "constraints": [
            "Must be buildable as a digital file/template/dashboard in this workspace.",
            "Must be sellable with Base Sepolia checkout metadata first.",
            "Do not invent source URLs; use URLs present in research_results.",
            "Do not return tool_name, arguments, command, structured_query, HTML, or markdown.",
        ],
        "research_results": compact_research,
    }
    for attempt in range(1, 3):
        try:
            reply = session.send(
                json.dumps(prompt, separators=(",", ":")),
                system=product_spec_prompt(),
                temperature=0.1,
                max_tokens=1536,
            )
        except Exception as exc:
            _append_log("strict_product_atf_failed", {"cycle": cycle, "attempt": attempt, "error": str(exc)})
            return {}
        try:
            spec = json.loads(_extract_json(reply))
            if isinstance(spec, dict) and not _looks_like_tool_response(spec):
                spec["sources"] = _top_sources(research_results)
                return spec
            _append_log(
                "strict_product_atf_tool_shaped_reply",
                {"cycle": cycle, "attempt": attempt, "spec": spec if isinstance(spec, dict) else str(type(spec))},
            )
            if hasattr(session, "report_outcome"):
                session.report_outcome(success=False, reason="product_spec_tool_shaped_reply")
        except Exception as exc:
            partial = _partial_product_spec_from_reply(reply, research_results)
            _append_log(
                "strict_product_atf_json_failed",
                {"cycle": cycle, "attempt": attempt, "error": str(exc), "partial": bool(partial), "reply_tail": reply[-1000:]},
            )
            if partial:
                return partial
    return {}


def _looks_like_tool_response(spec: Dict[str, Any]) -> bool:
    keys = {str(key).lower() for key in spec.keys()}
    if keys & {"tool_name", "arguments", "command", "structured_query", "tool_call_id"}:
        return True
    text = json.dumps(spec, ensure_ascii=True).lower()
    return any(marker in text for marker in ('"tool_name"', '"arguments"', '"command"', "structured_query", "bash"))


def _compact_research_for_model(research_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    compact = []
    for result in research_results:
        items = []
        candidates = [
            item for item in (result.get("results") or [])
            if isinstance(item, dict) and float(item.get("market_score") or 0.0) >= 0.20
        ] or [item for item in (result.get("results") or []) if isinstance(item, dict)]
        for item in candidates[:3]:
            if not isinstance(item, dict):
                continue
            items.append(
                {
                    "title": str(item.get("title") or "")[:120],
                    "url": str(item.get("url") or "")[:240],
                    "snippet": str(item.get("snippet") or "")[:220],
                    "market_score": item.get("market_score", 0),
                }
            )
        compact.append({"query": result.get("query"), "summary": str(result.get("summary") or "")[:600], "results": items})
    return compact


def _top_sources(research_results: List[Dict[str, Any]], limit: int = 4) -> List[Dict[str, str]]:
    scored_sources: List[Dict[str, Any]] = []
    for result in research_results:
        for item in result.get("results") or []:
            if isinstance(item, dict) and item.get("url"):
                scored_sources.append(
                    {
                        "title": str(item.get("title") or "Source"),
                        "url": str(item.get("url") or ""),
                        "snippet": str(item.get("snippet") or "")[:500],
                        "market_score": float(item.get("market_score") or 0.0),
                    }
                )
    scored_sources.sort(key=lambda item: item.get("market_score", 0.0), reverse=True)
    return [
        {"title": item["title"], "url": item["url"], "snippet": item.get("snippet", ""), "market_score": item["market_score"]}
        for item in scored_sources[:limit]
    ]


def _partial_product_spec_from_reply(reply: str, research_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    import re

    text = reply or ""

    def grab(key: str, default: str = "") -> str:
        match = re.search(rf'"{re.escape(key)}"\s*:\s*"([^"]*)', text)
        return match.group(1).strip() if match else default

    name = grab("name")
    summary = grab("summary")
    if not name and not summary:
        return {}
    slug = grab("slug") or _slug(name)
    return {
        "_partial": True,
        "name": name or slug.replace("-", " ").title(),
        "slug": slug,
        "summary": summary or "Partially generated ATF product concept; requires refinement.",
        "buyer": grab("buyer", "small business operators"),
        "pain_point": grab("pain_point", "manual workflow inefficiency"),
        "deliverable": grab("deliverable", "digital worksheet/template"),
        "price_usdc": grab("price_usdc", "9"),
        "confidence": 0.35,
        "sources": _top_sources(research_results),
        "next_improvement": "Rerun ATF spec generation with stricter compact JSON and enrich the partial product.",
    }


def _fallback_product_spec(*, cycle: int, research_results: List[Dict[str, Any]]) -> Dict[str, Any]:
    sources = _top_sources(research_results, limit=4)
    top = sources[0] if sources else {}
    top_title = str(top.get("title") or "").lower()
    top_url = str(top.get("url") or "")
    if any(term in top_title for term in ("lead", "crm", "sales")):
        name = "Small Business Lead Follow-Up Tracker"
        slug = f"lead-follow-up-tracker-{cycle}"
        summary = "A practical spreadsheet/template kit for tracking sales leads, follow-up dates, urgency, and next actions using small-business CRM and lead-tracker evidence."
        buyer = "small business owners, solo consultants, freelancers, and service businesses"
        pain_point = "Leads get lost when follow-up dates, urgency, source, and owner notes are scattered across email, spreadsheets, and memory."
        deliverable = "Excel/Google Sheets tracker, follow-up checklist, and import-ready CSV schema."
        next_improvement = "Add reminder automation instructions for Google Apps Script, Zapier, and Airtable."
    elif any(term in top_title for term in ("workflow", "automation", "process")):
        name = "Workflow Automation Gap Finder"
        slug = f"workflow-automation-gap-finder-{cycle}"
        summary = "A research-backed worksheet for identifying repeated manual workflows, estimating time loss, and ranking automation opportunities."
        buyer = "small business operators, consultants, and internal operations managers"
        pain_point = "Teams know repeated manual work is slowing them down, but they lack a simple scoring model for deciding what to automate first."
        deliverable = "Markdown/PDF worksheet, scoring rubric, and automation candidate checklist."
        next_improvement = "Add example Zapier/Make/n8n automation recipes for the top-ranked workflow categories."
    else:
        name = "Small Business Operations Template Pack"
        slug = f"operations-template-pack-{cycle}"
        summary = "A compact digital template pack derived from practical market evidence for improving recurring small-business operations."
        buyer = "small business owners and solo operators"
        pain_point = "Operators need lightweight, low-cost digital tools that solve recurring workflow problems without a heavy platform rollout."
        deliverable = "Markdown worksheet, spreadsheet schema, checklist, and setup guide."
        next_improvement = "Use the next ATF pass to narrow the niche and add automation-specific implementation steps."
    if top_url and top_url not in summary:
        summary = f"{summary} Primary evidence anchor: {top_url}"
    return {
        "name": name,
        "slug": slug,
        "summary": summary,
        "buyer": buyer,
        "pain_point": pain_point,
        "deliverable": deliverable,
        "price_usdc": "9",
        "confidence": 0.58,
        "sources": sources,
        "next_improvement": next_improvement,
    }


def _normalize_product_price(spec: Dict[str, Any]) -> Dict[str, Any]:
    if not isinstance(spec, dict):
        return spec
    raw = str(spec.get("price_usdc") or "9").strip()
    match = re.search(r"\d+(?:\.\d+)?", raw)
    try:
        value = float(match.group(0)) if match else 9.0
    except Exception:
        value = 9.0
    deliverable = str(spec.get("deliverable") or spec.get("summary") or "").lower()
    simple_markers = ("template", "worksheet", "spreadsheet", "checklist", "pdf", "notion", "excel")
    if any(marker in deliverable for marker in simple_markers):
        value = min(value, 49.0)
        value = max(value, 5.0)
    else:
        value = min(max(value, 9.0), 199.0)
    normalized = int(value) if value.is_integer() else round(value, 2)
    return {**spec, "price_usdc": str(normalized)}


def _product_spec_quality_error(spec: Dict[str, Any]) -> str:
    if not isinstance(spec, dict) or not spec:
        return "empty_spec"
    name = str(spec.get("name") or "").strip()
    slug = str(spec.get("slug") or "").strip().lower()
    summary = str(spec.get("summary") or "").strip()
    buyer = str(spec.get("buyer") or "").strip()
    deliverable = str(spec.get("deliverable") or "").strip()
    try:
        confidence = float(spec.get("confidence") or 0.0)
    except Exception:
        confidence = 0.0
    generic_names = {"digital product", "product", "untitled", "template", "dashboard"}
    if not name or name.lower() in generic_names:
        return "generic_or_missing_name"
    if not slug or re.match(r"^product-\d+$", slug):
        return "generic_or_missing_slug"
    if len(summary) < 40:
        return "summary_too_thin"
    if len(buyer) < 10:
        return "buyer_too_thin"
    if len(deliverable) < 10:
        return "deliverable_too_thin"
    if confidence < 0.50:
        return "confidence_too_low"
    sources = spec.get("sources") if isinstance(spec.get("sources"), list) else []
    if not sources:
        return "missing_sources"
    strong = [
        source for source in sources
        if isinstance(source, dict) and float(source.get("market_score") or 0.0) >= 0.30
    ]
    if len(strong) < 2:
        return "sources_too_weak"
    pain_point = str(spec.get("pain_point") or "").strip()
    if len(pain_point) < 40:
        return "pain_point_too_thin"
    source_signal = " ".join(str(source.get("snippet") or "") for source in strong).lower()
    if not any(term in source_signal for term in ("manual", "problem", "frustrat", "wish", "hours", "repet", "pain", "workaround", "difficult")):
        return "sources_do_not_contain_observed_pain_signal"
    return ""


def _product_markdown(spec: Dict[str, Any], research_results: List[Dict[str, Any]]) -> str:
    sources = spec.get("sources") if isinstance(spec.get("sources"), list) else []
    source_lines = "\n".join(
        f"- score={s.get('market_score', 'n/a')} {s.get('title', 'Source')}: {s.get('url', '')}"
        for s in sources[:8]
        if isinstance(s, dict)
    )
    evidence_quality = _market_evidence_quality(research_results)
    useful_research = []
    for result in research_results:
        if not isinstance(result, dict):
            continue
        max_score = max(
            (float(item.get("market_score") or 0.0) for item in result.get("results") or [] if isinstance(item, dict)),
            default=0.0,
        )
        if max_score >= 0.30 or str(result.get("query") or "").startswith("curated:"):
            useful_research.append(result)
    research_lines = "\n".join(
        f"- {r.get('query')}: {str(r.get('summary') or '')[:500]}" for r in useful_research[:8]
    )
    return f"""# {spec.get('name') or 'Digital Product'}

## Summary

{spec.get('summary') or ''}

## Buyer

{spec.get('buyer') or ''}

## Pain point

{spec.get('pain_point') or ''}

## Deliverable

{spec.get('deliverable') or ''}

## Price

{spec.get('price_usdc') or '9'} USDC on Base Sepolia while checkout is in testnet validation.

## Research notes

Evidence quality: `{json.dumps(evidence_quality)}`

{research_lines or '- No research summary available.'}

## Sources

{source_lines or '- Source capture pending.'}

## Next improvement

{spec.get('next_improvement') or 'Improve product details and checkout UX.'}
"""


def _validate_product_workspace(root: Path) -> Dict[str, Any]:
    checks: List[str] = []
    try:
        catalog = json.loads((root / "site" / "products.json").read_text(encoding="utf-8"))
        assert isinstance(catalog.get("products"), list)
        checks.append("catalog_json_valid")
    except Exception as exc:
        return {"passed": False, "checks": checks, "error": f"catalog_invalid:{exc}"}
    if not (root / "site" / "index.html").exists():
        return {"passed": False, "checks": checks, "error": "missing_index_html"}
    checks.append("storefront_exists")
    if not (root / "base" / "checkout.contract.json").exists():
        return {"passed": False, "checks": checks, "error": "missing_base_checkout_contract"}
    checks.append("base_checkout_contract_exists")
    return {"passed": True, "checks": checks}


def _extract_json(text: str) -> str:
    raw = (text or "").strip()
    if raw.startswith("```"):
        parts = raw.split("```")
        for part in parts:
            cleaned = part.strip()
            if cleaned.lower().startswith("json"):
                return cleaned.split("\n", 1)[1] if "\n" in cleaned else "{}"
            if cleaned.startswith("{"):
                return cleaned
    start = raw.find("{")
    end = raw.rfind("}")
    if start >= 0 and end > start:
        return raw[start : end + 1]
    return raw


def _slug(value: str) -> str:
    import re

    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "digital-product"


def _record_cycle_artifact(project_id: Optional[str], cycle: int, state: Dict[str, Any], output: str) -> None:
    if not project_id:
        return
    try:
        project = BrandProject.objects.filter(id=project_id).first()
        if not project:
            return
        run = DeliveryRun.objects.create(
            id=uuid.uuid4(),
            project=project,
            prompt=f"BrandDozer product loop cycle {cycle}",
            mode="existing",
            status="complete" if state.get("status") == "completed" else "running" if state.get("status") == "in_progress" else "error",
            phase="product_loop",
            iteration=cycle,
            acceptance_required=False,
            context={"product_loop": True, "state_path": str(STATE_PATH), "workspace": state.get("workspace")},
            error=state.get("error", ""),
            started_at=timezone.now(),
            completed_at=timezone.now(),
        )
        DeliveryArtifact.objects.create(
            project=project,
            run=run,
            kind="completion_report",
            title=f"Product loop cycle {cycle}",
            content=json.dumps({"state": state, "output": output}, indent=2),
            path=str(STATE_PATH),
        )
    except Exception:
        return
